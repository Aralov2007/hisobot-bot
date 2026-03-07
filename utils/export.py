import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from typing import List, Dict
from database.models import Report
from datetime import datetime
import os

OUTPUT_DIR = "exports"
os.makedirs(OUTPUT_DIR, exist_ok=True)


async def export_to_excel(reports: List[Report], stats: Dict) -> str:
    wb = openpyxl.Workbook()

    # === SHEET 1: Reports ===
    ws1 = wb.active
    ws1.title = "Hisobotlar"

    # Header style
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["#", "Sana", "Xodim", "Ish turi", "Materiallar", "Izoh", "Holat"]
    col_widths = [5, 15, 20, 20, 40, 25, 12]

    ws1.row_dimensions[1].height = 25
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    status_text = {"pending": "Kutmoqda", "approved": "Tasdiqlandi", "rejected": "Rad etildi"}
    alt_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")

    for row_idx, report in enumerate(reports, 2):
        materials_str = "; ".join([
            f"{(rm.material.name if rm.material else rm.custom_name) or 'N/A'}: {rm.quantity} {rm.unit}"
            for rm in report.materials
        ])

        row_data = [
            row_idx - 1,
            report.report_date.strftime("%d.%m.%Y"),
            report.user.full_name if report.user else "N/A",
            report.work_type,
            materials_str,
            report.note or "",
            status_text.get(report.status.value, report.status.value)
        ]

        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill

        ws1.row_dimensions[row_idx].height = 40

    # === SHEET 2: Material Stats ===
    ws2 = wb.create_sheet("Material statistikasi")
    ws2.cell(row=1, column=1, value="Material nomi").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="Jami miqdor").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.cell(row=1, column=3, value="Birlik").font = header_font
    ws2.cell(row=1, column=3).fill = header_fill
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 10

    for row_idx, (name, data) in enumerate(sorted(stats.items(), key=lambda x: x[1]["total"], reverse=True), 2):
        ws2.cell(row=row_idx, column=1, value=name)
        ws2.cell(row=row_idx, column=2, value=data["total"])
        ws2.cell(row=row_idx, column=3, value=data["unit"])
        if row_idx % 2 == 0:
            for col in range(1, 4):
                ws2.cell(row=row_idx, column=col).fill = alt_fill

    # Title
    now = datetime.now()
    filename = f"{OUTPUT_DIR}/hisobot_{now.strftime('%Y_%m')}.xlsx"
    wb.save(filename)
    return filename


async def export_to_pdf(reports: List[Report], stats: Dict) -> str:
    now = datetime.now()
    filename = f"{OUTPUT_DIR}/hisobot_{now.strftime('%Y_%m')}.pdf"

    doc = SimpleDocTemplate(filename, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    story = []

    # Title
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=16, spaceAfter=12,
                                  textColor=colors.HexColor('#2E86AB'))
    story.append(Paragraph(f"Oylik Hisobot — {now.strftime('%B %Y')}", title_style))
    story.append(Paragraph(f"Jami hisobotlar: {len(reports)} ta", styles['Normal']))
    story.append(Spacer(1, 0.5*cm))

    # Reports table
    table_data = [["#", "Sana", "Xodim", "Ish turi", "Holat"]]
    status_text = {"pending": "Kutmoqda", "approved": "Tasdiqlandi", "rejected": "Rad etildi"}

    for i, r in enumerate(reports, 1):
        table_data.append([
            str(i),
            r.report_date.strftime("%d.%m.%Y"),
            (r.user.full_name if r.user else "N/A")[:20],
            r.work_type[:20],
            status_text.get(r.status.value, r.status.value)
        ])

    col_widths = [1*cm, 2.5*cm, 4*cm, 4*cm, 2.5*cm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86AB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F8FF')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    # Materials stats
    story.append(Paragraph("Material Statistikasi", title_style))
    mat_data = [["Material", "Miqdor", "Birlik"]]
    for name, data in sorted(stats.items(), key=lambda x: x[1]["total"], reverse=True):
        qty = data["total"]
        qty_str = str(int(qty)) if qty == int(qty) else f"{qty:.1f}"
        mat_data.append([name[:30], qty_str, data["unit"]])

    mat_col_widths = [8*cm, 3*cm, 3*cm]
    t2 = Table(mat_data, colWidths=mat_col_widths, repeatRows=1)
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#A23B72')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFF0F8')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(t2)

    # Footer
    story.append(Spacer(1, 1*cm))
    footer_text = f"Hisobot yaratildi: {now.strftime('%d.%m.%Y %H:%M')}"
    story.append(Paragraph(footer_text, styles['Normal']))

    doc.build(story)
    return filename
